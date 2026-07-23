import os
import sys
import json
import ast
import time
import pandas as pd
import numpy as np
import joblib
import xgboost as xgb
import torch
import torch.nn as nn
import torch.nn.functional as F
from torch_geometric.data import Data
from torch_geometric.nn import SAGEConv, GATConv
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, roc_auc_score, confusion_matrix
from sklearn.ensemble import IsolationForest
from sklearn.neighbors import NearestNeighbors
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.append('backend')
sys.path.append('backend/playground')

import eda_utils
import gnn_utils
from scaler_utils import TriChannelScaler

class GNNClassifier(torch.nn.Module):
    def __init__(self, input_dim, hidden_dim, num_classes, num_layers=2, dropout=0.3, arch='sage'):
        super().__init__()
        self.dropout = dropout
        self.arch = arch
        self.layers = torch.nn.ModuleList()
        
        # Input layer
        if arch == 'sage':
            self.layers.append(SAGEConv(input_dim, hidden_dim))
        elif arch == 'gat':
            self.layers.append(GATConv(input_dim, hidden_dim, heads=1))
            
        # Hidden layers
        for _ in range(num_layers - 2):
            if arch == 'sage':
                self.layers.append(SAGEConv(hidden_dim, hidden_dim))
            elif arch == 'gat':
                self.layers.append(GATConv(hidden_dim, hidden_dim, heads=1))
                
        # Output layer
        if num_layers > 1:
            if arch == 'sage':
                self.layers.append(SAGEConv(hidden_dim, num_classes))
            elif arch == 'gat':
                self.layers.append(GATConv(hidden_dim, num_classes, heads=1))
        else:
            pass

    def forward(self, x, edge_index):
        for i, layer in enumerate(self.layers):
            x = layer(x, edge_index)
            if i < len(self.layers) - 1:
                x = F.relu(x)
                x = F.dropout(x, p=self.dropout, training=self.training)
        return x

def build_graph_fast(df, feature_cols, strategy='hybrid', k=5, delta_t_seconds=10):
    df = df.reset_index(drop=True)
    x = torch.tensor(df[feature_cols].values, dtype=torch.float)
    y = torch.tensor(df['Label_Encoded'].values, dtype=torch.long)
    
    src_nodes_list = []
    dst_nodes_list = []
    
    if 'knn' in strategy or 'hybrid' in strategy:
        if 'Protocol' in df.columns:
            protocols = df['Protocol'].unique()
            for p in protocols:
                indices = df.index[df['Protocol'] == p].values
                if len(indices) < k + 1:
                    continue
                subset = df.iloc[indices][feature_cols].values
                nbrs = NearestNeighbors(n_neighbors=k+1, algorithm='ball_tree', n_jobs=-1).fit(subset)
                distances, knn_indices = nbrs.kneighbors(subset)
                neighbors_global = indices[knn_indices[:, 1:]]
                srcs = np.repeat(indices, k)
                dsts = neighbors_global.flatten()
                src_nodes_list.extend(srcs)
                dst_nodes_list.extend(dsts)
                
    if 'src_ip' in strategy or 'hybrid' in strategy:
        valid_time = df['Timestamp'].notna()
        grouped = df[valid_time].groupby('Src IP')
        src_nodes = []
        dst_nodes = []
        for ip, group in grouped:
            if len(group) < 2:
                continue
            indices = group.index.values
            times = group['Timestamp'].values
            time_diffs = (times[1:] - times[:-1]) / np.timedelta64(1, 's')
            valid_links = time_diffs <= delta_t_seconds
            from_idx = indices[:-1][valid_links]
            to_idx = indices[1:][valid_links]
            src_nodes.extend(from_idx)
            dst_nodes.extend(to_idx)
        src_nodes_list.extend(src_nodes)
        dst_nodes_list.extend(dst_nodes)
        
    if not src_nodes_list:
        edge_index = torch.empty((2, 0), dtype=torch.long)
    else:
        edge_index = torch.tensor([src_nodes_list, dst_nodes_list], dtype=torch.long)
        edge_index = torch.unique(edge_index, dim=1)
    return Data(x=x, y=y, edge_index=edge_index)

def evaluate_gnn_local(model, data, mask, normal_idx, is_binary):
    model.eval()
    with torch.no_grad():
        logits = model(data.x, data.edge_index)
        preds = logits.argmax(dim=1)
        
    y_true = data.y[mask].cpu().numpy()
    y_pred = preds[mask].cpu().numpy()
    
    acc = accuracy_score(y_true, y_pred)
    
    if is_binary:
        y_true_bin = y_true
        y_pred_bin = y_pred
    else:
        y_true_bin = (y_true != normal_idx).astype(int)
        y_pred_bin = (y_pred != normal_idx).astype(int)
        
    rec = recall_score(y_true_bin, y_pred_bin, zero_division=0)
    
    # FPR (Normal)
    tn, fp, fn, tp = confusion_matrix(y_true_bin, y_pred_bin, labels=[0, 1]).ravel()
    fpr = fp / (fp + tn) if (fp + tn) > 0 else 0
    
    if is_binary:
        f1 = f1_score(y_true, y_pred, zero_division=0)
        prec = precision_score(y_true, y_pred, zero_division=0)
    else:
        f1 = f1_score(y_true, y_pred, average='macro', zero_division=0)
        prec = precision_score(y_true, y_pred, average='macro', zero_division=0)
        
    return acc, prec, rec, fpr, f1

def evaluate_xgb_models_grid():
    print("[*] Running XGBoost grid search (4 splits x 3 seeds)...")
    df = pd.read_csv('backend/playground/checkpoints/ctgan/balanced_20.csv')
    y = (df['Label'] != 'Normal').astype(int)
    X = df.drop(columns=['Label'])

    scaler = joblib.load('backend/playground/scalers/benign_robust_scaler.pkl')
    scaled_cols = list(scaler.feature_names_in_)

    xgb_features = [
        'Dst Port', 'Protocol', 'Flow Duration', 'Tot Fwd Pkts', 'Bwd Pkt Len Max', 
        'Bwd Pkt Len Min', 'Flow Pkts/s', 'Flow IAT Std', 'Flow IAT Max', 'Flow IAT Min', 
        'Fwd IAT Tot', 'Fwd IAT Mean', 'Fwd IAT Max', 'Bwd IAT Mean', 'Bwd IAT Min', 
        'Fwd Header Len', 'Fwd Pkts/s', 'Pkt Len Max', 'Pkt Len Mean', 'Init Bwd Win Byts'
    ]

    splits = [(0.70, 0.15, 0.15), (0.60, 0.20, 0.20), (0.50, 0.25, 0.25), (0.40, 0.30, 0.30)]
    seeds = [42, 52, 62]
    
    xgb_results = []
    for train_r, val_r, test_r in splits:
        for seed in seeds:
            # Replicate get_golden_split logic
            X_temp, X_test, y_temp, y_test = train_test_split(
                X, y, test_size=test_r, random_state=seed, stratify=y
            )
            relative_val = val_r / (1.0 - test_r)
            X_train, X_val, y_train, y_val = train_test_split(
                X_temp, y_temp, test_size=relative_val, random_state=seed, stratify=y_temp
            )
            
            X_train_scaled = X_train.copy()
            X_val_scaled = X_val.copy()
            X_test_scaled = X_test.copy()
            
            X_train_scaled[scaled_cols] = scaler.transform(X_train_scaled[scaled_cols])
            X_val_scaled[scaled_cols] = scaler.transform(X_val_scaled[scaled_cols])
            X_test_scaled[scaled_cols] = scaler.transform(X_test_scaled[scaled_cols])
            
            model = xgb.XGBClassifier(
                device="cuda",              
                tree_method="hist",         
                objective="binary:logistic",
                n_estimators=400,           
                learning_rate=0.1,          
                max_depth=6,
                subsample=0.8,
                colsample_bytree=0.8,
                eval_metric=["logloss", "auc"],
                random_state=seed,
                early_stopping_rounds=20
            )
            
            X_tr_in = X_train_scaled[xgb_features]
            X_val_in = X_val_scaled[xgb_features]
            X_te_in = X_test_scaled[xgb_features]
            
            model.fit(
                X_tr_in, y_train,
                eval_set=[(X_tr_in, y_train), (X_val_in, y_val)],
                verbose=False
            )
            
            preds = model.predict(X_te_in)
            probs = model.predict_proba(X_te_in)[:, 1]
            
            acc = accuracy_score(y_test, preds)
            prec = precision_score(y_test, preds, zero_division=0)
            rec = recall_score(y_test, preds, zero_division=0)
            f1 = f1_score(y_test, preds, zero_division=0)
            auc = roc_auc_score(y_test, probs)
            
            xgb_results.append({
                'Split (Train/Val/Test)': f'{int(train_r*100)}/{int(val_r*100)}/{int(test_r*100)}',
                'Seed': seed,
                'Train Size': len(y_train),
                'Val Size': len(y_val),
                'Test Size': len(y_test),
                'n_estimators': 400,
                'learning_rate': 0.1,
                'max_depth': 6,
                'Accuracy': float(acc),
                'Precision': float(prec),
                'Recall': float(rec),
                'F1-Score': float(f1),
                'ROC-AUC': float(auc)
            })
            print(f"XGB split {train_r}/{val_r}/{test_r} seed {seed} complete. F1: {f1:.4f}")
            
    return pd.DataFrame(xgb_results)

def evaluate_if_models_grid():
    print("[*] Running Isolation Forest grid search (excluding retrained models)...")
    df = pd.read_csv('backend/02ctgan/balanced_15features.csv')
    y_true = (df['Label'] != 'Normal').astype(int)
    X_raw = df.drop(columns=['Label'])
    
    scaler = joblib.load('backend/scalers/trichannel_scaler.pkl')
    X_scaled = scaler.transform(X_raw)
    
    X_train, X_test, y_train, y_test = train_test_split(
        X_scaled, y_true, test_size=0.3, stratify=y_true, random_state=42
    )
    
    X_train_benign = X_train[y_train == 0]
    
    if_results = []
    
    sn_baseline = joblib.load('backend/models/safetynet/safety_net_v1.pkl')
    base_preds = sn_baseline.predict(X_test)
    if_results.append({
        'Model Source': 'safety_net_v1.pkl (Pre-trained Baseline)',
        'n_estimators': 100,
        'contamination': 0.01,
        'Accuracy': float(accuracy_score(y_test, base_preds)),
        'Precision': float(precision_score(y_test, base_preds, zero_division=0)),
        'Recall': float(recall_score(y_test, base_preds, zero_division=0)),
        'F1-Score': float(f1_score(y_test, base_preds, zero_division=0))
    })
    
    contaminations = [0.001, 0.005, 0.01, 0.02, 0.05]
    n_estimators_list = [50, 100, 200]
    
    for n_est in n_estimators_list:
        for cont in contaminations:
            model = IsolationForest(
                n_estimators=n_est,
                contamination=cont,
                random_state=42,
                n_jobs=-1
            )
            model.fit(X_train_benign)
            
            raw_preds = model.predict(X_test)
            preds = np.where(raw_preds == -1, 1, 0)
            
            acc = accuracy_score(y_test, preds)
            prec = precision_score(y_test, preds, zero_division=0)
            rec = recall_score(y_test, preds, zero_division=0)
            f1 = f1_score(y_test, preds, zero_division=0)
            
            if_results.append({
                'Model Source': 'Grid Search Swept Model',
                'n_estimators': n_est,
                'contamination': cont,
                'Accuracy': float(acc),
                'Precision': float(prec),
                'Recall': float(rec),
                'F1-Score': float(f1)
            })
            print(f"IF n_est={n_est} cont={cont} complete. F1: {f1:.4f}")
            
    return pd.DataFrame(if_results)

def run_gnn_sweep(df_raw, le, normal_idx, device):
    print("[*] Starting GNN Parameter Sweep (864 configurations)...")
    
    df_ref_15 = pd.read_csv('backend/01eda/cleaned_data15.csv', nrows=1)
    feature_cols_15 = [c for c in df_ref_15.columns if c != 'Label']
    
    df_ref_51 = pd.read_csv('backend/01eda/cleaned_data52.csv', nrows=1)
    feature_cols_51 = [c for c in df_ref_51.columns if c != 'Label']
    
    # Downsample df_raw for training sweep (take every 11th row -> 31,262 rows)
    df_sampled = df_raw.iloc[::11].copy().reset_index(drop=True)
    print(f"  Downsampled training dataset size: {len(df_sampled)} rows")
    
    from sklearn.preprocessing import StandardScaler, RobustScaler
    
    source_scalers = {}
    for n_feat in [15, 51]:
        feats = feature_cols_15 if n_feat == 15 else feature_cols_51
        df_feats = df_sampled[feats].copy()
        
        # Standard
        std = StandardScaler().fit(df_feats)
        source_scalers[(n_feat, 'standard')] = std
        
        # Robust
        rob = RobustScaler().fit(df_feats)
        source_scalers[(n_feat, 'robust')] = rob
        
        # Tri-Channel
        tc = TriChannelScaler(benign_label=normal_idx)
        tc.fit(df_feats, df_sampled['Label_Encoded'])
        source_scalers[(n_feat, 'tri-channel')] = tc
        
    # Build graphs cache
    graphs = {}
    for n_feat in [15, 51]:
        feats = feature_cols_15 if n_feat == 15 else feature_cols_51
        df_feats = df_sampled.dropna(subset=feats).copy().reset_index(drop=True)
        
        for scaler_type in ['standard', 'robust', 'tri-channel']:
            df_scaled = df_feats.copy()
            scaler_obj = source_scalers[(n_feat, scaler_type)]
            
            if scaler_type == 'tri-channel':
                df_tc = scaler_obj.transform(df_scaled[feats])
                feats_to_use = df_tc.columns.tolist()
                meta_cols = ['Src IP', 'Dst IP', 'Dst Port', 'Protocol', 'Timestamp', 'Label_Encoded']
                df_scaled = pd.concat([df_tc, df_scaled[meta_cols]], axis=1)
            else:
                df_scaled[feats] = scaler_obj.transform(df_scaled[feats])
                feats_to_use = feats
                
            for strategy in ['knn_protocol', 'src_ip_temporal', 'hybrid']:
                graph = build_graph_fast(df_scaled, feats_to_use, strategy=strategy)
                graphs[(n_feat, scaler_type, strategy)] = (graph.to(device), feats_to_use)
                
    # Sweep Loops
    scalers = ['standard', 'robust', 'tri-channel']
    tasks = ['binary', 'multiclass']
    strategies = ['knn_protocol', 'src_ip_temporal', 'hybrid']
    architectures = ['sage', 'gat']
    splits = [(0.70, 0.15, 0.15), (0.80, 0.10, 0.10), (0.60, 0.20, 0.20)]
    n_features_list = [15, 51]
    hidden_dims = [64, 128]
    num_layers_list = [2, 3]
    
    gnn_results = []
    run_idx = 0
    t_start = time.time()
    
    for scaler_type in scalers:
        for task in tasks:
            for strategy in strategies:
                for arch in architectures:
                    for split in splits:
                        for n_feat in n_features_list:
                            for hidden_dim in hidden_dims:
                                for num_layers in num_layers_list:
                                    run_idx += 1
                                    
                                    graph_orig, feats_used = graphs[(n_feat, scaler_type, strategy)]
                                    data = graph_orig.clone()
                                    
                                    if task == 'binary':
                                        data.y = (graph_orig.y != normal_idx).long()
                                        num_classes = 2
                                        is_binary = True
                                    else:
                                        num_classes = len(le.classes_)
                                        is_binary = False
                                        
                                    num_nodes = data.num_nodes
                                    indices = np.arange(num_nodes)
                                    y_cpu = data.y.cpu().numpy()
                                    
                                    train_ratio, val_ratio, test_ratio = split
                                    try:
                                        train_idx, temp_idx = train_test_split(
                                            indices, train_size=train_ratio, stratify=y_cpu, random_state=42
                                        )
                                        val_ratio_adj = val_ratio / (val_ratio + test_ratio)
                                        val_idx, test_idx = train_test_split(
                                            temp_idx, train_size=val_ratio_adj, stratify=y_cpu[temp_idx], random_state=42
                                        )
                                    except ValueError:
                                        train_idx, temp_idx = train_test_split(
                                            indices, train_size=train_ratio, random_state=42
                                        )
                                        val_ratio_adj = val_ratio / (val_ratio + test_ratio)
                                        val_idx, test_idx = train_test_split(
                                            temp_idx, train_size=val_ratio_adj, random_state=42
                                        )
                                        
                                    train_mask = torch.zeros(num_nodes, dtype=torch.bool).to(device)
                                    val_mask = torch.zeros(num_nodes, dtype=torch.bool).to(device)
                                    test_mask = torch.zeros(num_nodes, dtype=torch.bool).to(device)
                                    
                                    train_mask[train_idx] = True
                                    val_mask[val_idx] = True
                                    test_mask[test_idx] = True
                                    
                                    # Class Weights
                                    y_train = y_cpu[train_idx]
                                    classes_in_train = np.unique(y_train)
                                    class_weights = torch.tensor(
                                        [len(y_train) / (len(classes_in_train) * np.sum(y_train == c)) for c in range(num_classes)],
                                        dtype=torch.float
                                    ).to(device)
                                    
                                    dropout = 0.3 if hidden_dim == 64 else 0.5
                                    model = GNNClassifier(
                                        input_dim=len(feats_used),
                                        hidden_dim=hidden_dim,
                                        num_classes=num_classes,
                                        num_layers=num_layers,
                                        dropout=dropout,
                                        arch=arch
                                    ).to(device)
                                    
                                    t0_fit = time.time()
                                    optimizer = torch.optim.Adam(model.parameters(), lr=0.01)
                                    criterion = torch.nn.CrossEntropyLoss(weight=class_weights)
                                    
                                    for epoch in range(10):
                                        model.train()
                                        optimizer.zero_grad()
                                        out = model(data.x, data.edge_index)
                                        loss = criterion(out[train_mask], data.y[train_mask])
                                        loss.backward()
                                        optimizer.step()
                                        
                                    fit_time = time.time() - t0_fit
                                    acc, prec, rec, fpr, f1 = evaluate_gnn_local(model, data, test_mask, normal_idx, is_binary)
                                    
                                    gnn_results.append({
                                        'Scaler': scaler_type,
                                        'Task': task,
                                        'Strategy': strategy,
                                        'Architecture': arch,
                                        'Split': f'{int(train_ratio*100)}/{int(val_ratio*100)}/{int(test_ratio*100)}',
                                        'N Features': n_feat,
                                        'Hidden Dim': hidden_dim,
                                        'Number of Layers': num_layers,
                                        'Accuracy': acc,
                                        'Precision': prec,
                                        'Recall': rec,
                                        'FPR': fpr,
                                        'F1-Score': f1,
                                        'Training Time (s)': fit_time
                                    })
                                    
                                    if run_idx % 100 == 0:
                                        print(f"  [GNN Sweep] Run {run_idx}/864 | Scaler={scaler_type} | Task={task} | F1={f1:.4f}", flush=True)
                                        
    print(f"[+] GNN Parameter Sweep complete in {time.time() - t_start:.2f} seconds.")
    return pd.DataFrame(gnn_results), source_scalers, feature_cols_15, feature_cols_51, df_sampled

def train_source_model(scaler_type, strategy, arch, n_feat, hidden_dim, num_layers, task, df_sampled, source_scalers, feats, normal_idx, le, device):
    df_scaled = df_sampled.copy()
    scaler_obj = source_scalers[(n_feat, scaler_type)]
    if scaler_type == 'tri-channel':
        df_tc = scaler_obj.transform(df_scaled[feats])
        feats_used = df_tc.columns.tolist()
        df_scaled = pd.concat([df_tc, df_scaled[['Src IP', 'Dst IP', 'Dst Port', 'Protocol', 'Timestamp', 'Label_Encoded']]], axis=1)
    else:
        df_scaled[feats] = scaler_obj.transform(df_scaled[feats])
        feats_used = feats
        
    graph_source = build_graph_fast(df_scaled, feats_used, strategy=strategy).to(device)
    
    if task == 'binary':
        graph_source.y = (graph_source.y != normal_idx).long()
        num_classes = 2
    else:
        num_classes = len(le.classes_)
        
    model = GNNClassifier(
        input_dim=len(feats_used),
        hidden_dim=hidden_dim,
        num_classes=num_classes,
        num_layers=num_layers,
        dropout=0.3 if hidden_dim == 64 else 0.5,
        arch=arch
    ).to(device)
    
    optimizer = torch.optim.Adam(model.parameters(), lr=0.01)
    criterion = torch.nn.CrossEntropyLoss()
    
    for epoch in range(15):
        model.train()
        optimizer.zero_grad()
        out = model(graph_source.x, graph_source.edge_index)
        loss = criterion(out, graph_source.y)
        loss.backward()
        optimizer.step()
        
    return model, feats_used

# Column mapping from raw columns to normalized names
COL_MAP = {
    'Source IP': 'Src IP', 'Source Port': 'Src Port', 'Destination IP': 'Dst IP',
    'Destination Port': 'Dst Port', 'Protocol': 'Protocol', 'Timestamp': 'Timestamp',
    'Flow Duration': 'Flow Duration', 'Total Fwd Packets': 'Tot Fwd Pkts',
    'Total Backward Packets': 'Tot Bwd Pkts', 'Total Length of Fwd Packets': 'TotLen Fwd Pkts',
    'Total Length of Bwd Packets': 'TotLen Bwd Pkts', 'Fwd Packet Length Max': 'Fwd Pkt Len Max',
    'Fwd Packet Length Min': 'Fwd Pkt Len Min', 'Fwd Packet Length Mean': 'Fwd Pkt Len Mean',
    'Fwd Packet Length Std': 'Fwd Pkt Len Std', 'Bwd Packet Length Max': 'Bwd Pkt Len Max',
    'Bwd Packet Length Min': 'Bwd Pkt Len Min', 'Bwd Packet Length Mean': 'Bwd Pkt Len Mean',
    'Bwd Packet Length Std': 'Bwd Pkt Len Std', 'Flow Bytes/s': 'Flow Byts/s',
    'Flow Packets/s': 'Flow Pkts/s', 'Flow IAT Mean': 'Flow IAT Mean',
    'Flow IAT Std': 'Flow IAT Std', 'Flow IAT Max': 'Flow IAT Max',
    'Flow IAT Min': 'Flow IAT Min', 'Fwd IAT Total': 'Fwd IAT Tot',
    'Fwd IAT Mean': 'Fwd IAT Mean', 'Fwd IAT Std': 'Fwd IAT Std',
    'Fwd IAT Max': 'Fwd IAT Max', 'Fwd IAT Min': 'Fwd IAT Min',
    'Bwd IAT Total': 'Bwd IAT Tot', 'Bwd IAT Mean': 'Bwd IAT Mean',
    'Bwd IAT Std': 'Bwd IAT Std', 'Bwd IAT Max': 'Bwd IAT Max',
    'Bwd IAT Min': 'Bwd IAT Min', 'Bwd PSH Flags': 'Bwd PSH Flags',
    'Bwd URG Flags': 'Bwd URG Flags', 'Fwd Header Length': 'Fwd Header Len',
    'Bwd Header Length': 'Bwd Header Len', 'Fwd Packets/s': 'Fwd Pkts/s',
    'Bwd Packets/s': 'Bwd Pkts/s', 'Min Packet Length': 'Pkt Len Min',
    'Max Packet Length': 'Pkt Len Max', 'Packet Length Mean': 'Pkt Len Mean',
    'Packet Length Std': 'Pkt Len Std', 'Packet Length Variance': 'Pkt Len Var',
    'FIN Flag Count': 'FIN Flag Cnt', 'SYN Flag Count': 'SYN Flag Cnt',
    'RST Flag Count': 'RST Flag Cnt', 'PSH Flag Count': 'PSH Flag Cnt',
    'ACK Flag Count': 'ACK Flag Cnt', 'URG Flag Count': 'URG Flag Cnt',
    'ECE Flag Count': 'ECE Flag Cnt', 'Down/Up Ratio': 'Down/Up Ratio',
    'Average Packet Size': 'Pkt Size Avg', 'Avg Fwd Segment Size': 'Fwd Seg Size Avg',
    'Avg Bwd Segment Size': 'Bwd Seg Size Avg', 'Fwd Avg Bytes/Bulk': 'Fwd Byts/b Avg',
    'Fwd Avg Packets/Bulk': 'Fwd Pkts/b Avg', 'Fwd Avg Bulk Rate': 'Fwd Blk Rate Avg',
    'Bwd Avg Bytes/Bulk': 'Bwd Byts/b Avg', 'Bwd Avg Packets/Bulk': 'Bwd Pkts/b Avg',
    'Bwd Avg Bulk Rate': 'Bwd Blk Rate Avg', 'Subflow Fwd Packets': 'Subflow Fwd Pkts',
    'Subflow Fwd Bytes': 'Subflow Fwd Byts', 'Subflow Bwd Packets': 'Subflow Bwd Pkts',
    'Subflow Bwd Bytes': 'Subflow Bwd Byts', 'Init_Win_bytes_forward': 'Init Fwd Win Byts',
    'Init_Win_bytes_backward': 'Init Bwd Win Byts', 'act_data_pkt_fwd': 'Fwd Act Data Pkts',
    'min_seg_size_forward': 'Fwd Seg Size Min', 'Active Mean': 'Active Mean',
    'Active Std': 'Active Std', 'Active Max': 'Active Max', 'Active Min': 'Active Min',
    'Idle Mean': 'Idle Mean', 'Idle Std': 'Idle Std', 'Idle Max': 'Idle Max',
    'Idle Min': 'Idle Min', 'Label': 'Label'
}

LABEL_MAP_DNS = {
    'BENIGN': 'Normal',
    'DrDoS_DNS': 'DDoS'
}

LABEL_MAP_FRIDAY = {
    'Benign': 'Normal',
    'DoS attacks-Hulk': 'DoS',
    'DoS attacks-SlowHTTPTest': 'DoS'
}

def load_and_preprocess_target_dataset(csv_path, dataset_type, feature_cols, nrows):
    df = pd.read_csv(csv_path, nrows=nrows, low_memory=False)
    df.columns = df.columns.str.strip()
    df.rename(columns=COL_MAP, inplace=True)
    if dataset_type == 'dns':
        df['Label'] = df['Label'].map(LABEL_MAP_DNS).fillna('Normal')
    elif dataset_type == 'friday':
        df['Label'] = df['Label'].map(LABEL_MAP_FRIDAY).fillna('Normal')
        
    if 'Src IP' not in df.columns:
        df['Src IP'] = '10.0.0.1'
    if 'Dst IP' not in df.columns:
        df['Dst IP'] = '10.0.0.2'
        
    for col in feature_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
            
    df.dropna(subset=feature_cols, inplace=True)
    df.replace([np.inf, -np.inf], np.nan, inplace=True)
    df.dropna(subset=feature_cols, inplace=True)
    return df

def run_ablation_study_local(best_models_df, df_sampled, le, normal_idx, source_scalers, feature_cols_15, feature_cols_51, device):
    import subprocess
    print("[*] Running run_ablation_study.py to execute the ablation study and select strategies...")
    subprocess.run(["backend/fypenv/bin/python3", "backend/gnn_compare/run_ablation_study.py"], check=True)
    
    # Load the results
    df_strategies = pd.read_csv("backend/gnn_compare/ablation_strategies.csv")
    df_results = pd.read_csv("backend/gnn_compare/ablation_study_results.csv")
    
    return df_strategies, df_results

def main():
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    print(f"Using Device: {device}", flush=True)
    
    # 1. Run XGB and IF grids
    xgb_df = evaluate_xgb_models_grid()
    if_df = evaluate_if_models_grid()
    
    # 2. GNN data loading
    print("Loading raw data to recover topology...", flush=True)
    filenames = ['metasploitable-2.csv', 'OVS.csv', 'Normal_data.csv']
    df_raw = eda_utils.load_and_concatenate_datasets('backend/datasets', filenames)
    df_raw['Label'] = (
        df_raw['Label']
        .astype(str)
        .str.replace(r"[\u200b\u200c\u200d\ufeff]", "", regex=True)
        .str.strip()
        .str.replace(r"\s+", " ", regex=True)
    )
    df_raw['Timestamp'] = pd.to_datetime(df_raw['Timestamp'], errors='coerce')
    df_raw = df_raw.sort_values('Timestamp').reset_index(drop=True)
    
    with open('backend/encoders/label_encoder.pkl', 'rb') as f:
        le = joblib.load(f)
        
    df_raw = df_raw[df_raw['Label'].isin(le.classes_)].reset_index(drop=True)
    df_raw['Label_Encoded'] = le.transform(df_raw['Label'])
    normal_idx = le.transform(['Normal'])[0]
    
    # 3. GNN Sweep (864 configurations)
    gnn_sweep_df, source_scalers, feature_cols_15, feature_cols_51, df_sampled = run_gnn_sweep(df_raw, le, normal_idx, device)
    
    # 4. GNN Model Selection (6 models)
    best_models_list = []
    for scaler in ['standard', 'robust', 'tri-channel']:
        for task in ['binary', 'multiclass']:
            df_sub = gnn_sweep_df[(gnn_sweep_df['Scaler'] == scaler) & (gnn_sweep_df['Task'] == task)]
            if not df_sub.empty:
                best_idx = df_sub['F1-Score'].idxmax()
                best_models_list.append(df_sub.loc[best_idx])
    best_models_df = pd.DataFrame(best_models_list)
    print("\n[+] Selected 6 Best GNN Models for Ablation Study:")
    print(best_models_df[['Scaler', 'Task', 'Strategy', 'Architecture', 'N Features', 'F1-Score']])
    
    # 5. Run GNN Ablation Study (36 configurations)
    df_strategies, df_results = run_ablation_study_local(best_models_df, df_sampled, le, normal_idx, source_scalers, feature_cols_15, feature_cols_51, device)
    
    # 6. Recompile Excel Report
    print("[*] Generating Workbook...")
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Define styles
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid') # Navy
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    title_font = Font(name='Calibri', size=16, bold=True, color='1F4E79')
    subtitle_font = Font(name='Calibri', size=11, italic=True, color='595959')
    bold_font = Font(name='Calibri', size=11, bold=True)
    normal_font = Font(name='Calibri', size=11)
    
    zebra_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    thin_border_side = Side(style='thin', color='D9D9D9')
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    def apply_table_styles(ws, start_row, df, border=cell_border):
        # Header
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = cell_border
        
        # Data
        for row_idx, row_data in enumerate(df.values):
            r = start_row + 1 + row_idx
            fill = zebra_fill if row_idx % 2 == 1 else white_fill
            for col_idx, val in enumerate(row_data):
                cell = ws.cell(row=r, column=col_idx + 1, value=val)
                cell.fill = fill
                cell.font = normal_font
                cell.border = border
                
                if isinstance(val, (int, np.integer)):
                    cell.alignment = align_right
                    cell.number_format = '#,##0'
                elif isinstance(val, (float, np.float64, np.float32)):
                    cell.alignment = align_right
                    cell.number_format = '0.0000'
                else:
                    if str(val).startswith('0.') or str(val).startswith('-0.'):
                        try:
                            fval = float(val)
                            cell.value = fval
                            cell.alignment = align_right
                            cell.number_format = '0.0000'
                        except:
                            cell.alignment = align_left
                    else:
                        cell.alignment = align_left
                        
    def autofit_columns(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                # Skip title rows and section header rows to prevent inflating column width
                if cell.row in [1, 2] or val_str.startswith("Dataset:") or val_str.startswith("Table "):
                    continue
                if cell.number_format and ('0.' in cell.number_format):
                    val_str = f"{cell.value:.4f}" if isinstance(cell.value, float) else val_str
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    def apply_styled_table(ws, start_row, df, headers, alignments, number_formats):
        # Header
        for col_idx, col_name in enumerate(headers):
            cell = ws.cell(row=start_row, column=col_idx + 1, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = cell_border
        
        # Data
        for row_idx, row_data in enumerate(df.values):
            r = start_row + 1 + row_idx
            fill = zebra_fill if row_idx % 2 == 1 else white_fill
            for col_idx, val in enumerate(row_data):
                cell = ws.cell(row=r, column=col_idx + 1)
                
                # Check if val is NaN or numeric or string
                if pd.isna(val):
                    cell.value = ""
                else:
                    cell.value = val
                    
                cell.fill = fill
                cell.font = normal_font
                cell.border = cell_border
                
                # Alignment
                if col_idx < len(alignments):
                    cell.alignment = alignments[col_idx]
                else:
                    cell.alignment = align_left
                
                # Format
                if col_idx < len(number_formats) and number_formats[col_idx] is not None:
                    if isinstance(val, (int, float, np.number)):
                        cell.number_format = number_formats[col_idx]
                    else:
                        try:
                            fval = float(val)
                            cell.value = fval
                            cell.number_format = number_formats[col_idx]
                        except:
                            pass

    # Sheet 1: XGB Models
    print("[*] Generating Sheet 1: XGB Models...")
    ws1 = wb.create_sheet(title="XGB Models")
    ws1.views.sheetView[0].showGridLines = True
    ws1.cell(row=1, column=1, value="XGBoost Classifier Parameters and Grid Search Results").font = title_font
    ws1.cell(row=2, column=1, value="Sweep results across 4 Train/Val/Test splits and 3 random seeds (12 configurations total)").font = subtitle_font
    for col_idx, col_name in enumerate(xgb_df.columns):
        ws1.cell(row=4, column=col_idx + 1, value=col_name)
    apply_table_styles(ws1, 4, xgb_df)
    autofit_columns(ws1)

    # Sheet 2: Isolation Forest
    print("[*] Generating Sheet 2: Isolation Forest...")
    ws2 = wb.create_sheet(title="Isolation Forest")
    ws2.views.sheetView[0].showGridLines = True
    ws2.cell(row=1, column=1, value="SafetyNet (Isolation Forest) Parameter Grid Search").font = title_font
    ws2.cell(row=2, column=1, value="Records different estimators and contamination parameters. Retrain models are excluded.").font = subtitle_font
    for col_idx, col_name in enumerate(if_df.columns):
        ws2.cell(row=4, column=col_idx + 1, value=col_name)
    apply_table_styles(ws2, 4, if_df)
    autofit_columns(ws2)

    # Sheet 3: GNN Models (864 rows)
    print("[*] Generating Sheet 3: GNN Models...")
    ws3 = wb.create_sheet(title="GNN Models")
    ws3.views.sheetView[0].showGridLines = True
    ws3.cell(row=1, column=1, value="GNN Parameters and Training Results").font = title_font
    ws3.cell(row=2, column=1, value="Unified grid search records of 864 runs varying Scaler, Task, Graph Strategy, Architecture, Splits, Features, Hidden Dim, and Depth.").font = subtitle_font
    for col_idx, col_name in enumerate(gnn_sweep_df.columns):
        ws3.cell(row=4, column=col_idx + 1, value=col_name)
    apply_table_styles(ws3, 4, gnn_sweep_df)
    autofit_columns(ws3)

    # Sheet 4: Ablation Report
    print("[*] Generating Sheet 4: Ablation Report...")
    ws4 = wb.create_sheet(title="Ablation Report")
    ws4.views.sheetView[0].showGridLines = True
    
    # Title & Subtitle
    ws4.cell(row=1, column=1, value="Adaptive GNN Ablation Study — Rescale vs Retrain Report").font = title_font
    ws4.cell(row=2, column=1, value="Comparative evaluation of GNN model adaptation on DNS and Friday target environments.").font = subtitle_font
    
    # Compute Key Insights first to check values
    # 1. Which dataset benefits more from rescale?
    m1_dns_rescale_imp = df_results[(df_results['dataset'] == 'dns') & (df_results['Model'] == 'MODEL1') & (df_results['Mode'] == 'Rescale')]['Improvement'].mean()
    m1_friday_rescale_imp = df_results[(df_results['dataset'] == 'friday') & (df_results['Model'] == 'MODEL1') & (df_results['Mode'] == 'Rescale')]['Improvement'].mean()
    
    if m1_dns_rescale_imp > m1_friday_rescale_imp:
        better_rescale_ds = f"DNS (average Model 1 F1 improvement: +{m1_dns_rescale_imp:.4f} vs Friday: {m1_friday_rescale_imp:+.4f})"
    else:
        better_rescale_ds = f"Friday (average Model 1 F1 improvement: +{m1_friday_rescale_imp:.4f} vs DNS: {m1_dns_rescale_imp:+.4f})"

    # 2. Which model collapses without adaptation?
    m_baseline = df_results[df_results['Mode'] == 'Baseline'].groupby(['Model', 'dataset'])['F1'].mean().reset_index()
    lowest_row = m_baseline.loc[m_baseline['F1'].idxmin()]
    collapsing_model = f"{lowest_row['Model']} on {lowest_row['dataset'].upper()} (F1 score: {lowest_row['F1']:.4f})"

    # 3. Does retraining consistently outperform rescale?
    df_res = df_results[df_results['Mode'] == 'Rescale'].set_index(['dataset', 'Model', 'Task'])
    df_ret = df_results[df_results['Mode'] == 'Retrain'].set_index(['dataset', 'Model', 'Task'])
    better_count = np.sum(df_ret['F1'] > df_res['F1'])
    total_count = len(df_res)
    does_outperform = "Yes" if better_count > (total_count / 2) else "No"
    outperform_str = f"{does_outperform} (Retraining was better in {better_count}/{total_count} model configurations)"

    # 4. When is rescale sufficient?
    close_configs = []
    for idx, row in df_res.iterrows():
        ret_f1 = df_ret.loc[idx, 'F1']
        res_f1 = row['F1']
        diff = ret_f1 - res_f1
        if diff <= 0.05:
            close_configs.append(f"{idx[1]} {idx[2]} on {idx[0].upper()}")
    
    if close_configs:
        rescale_sufficient = f"Rescale is sufficient in {len(close_configs)}/{total_count} configurations (F1 diff <= 0.05): {', '.join(close_configs)}"
    else:
        rescale_sufficient = "Rescale is rarely sufficient; retraining is needed to recover discriminative power."

    current_row = 4
    
    # Sub-styles for subheadings
    section_font = Font(name='Calibri', size=14, bold=True, color='1F4E79')
    table_title_font = Font(name='Calibri', size=11, bold=True)
    
    # ----------------------------------------------------
    # Dataset: DNS
    # ----------------------------------------------------
    ws4.cell(row=current_row, column=1, value="Dataset: DNS").font = section_font
    current_row += 2
    
    # Table A: Calibration Strategy Selection (DNS)
    ws4.cell(row=current_row, column=1, value="Table A: Calibration Strategy Selection (DNS)").font = table_title_font
    ws4.cell(row=current_row + 1, column=1, value="Evaluated on the first 20,000 flows. Objective: avoid extreme ratio collapse.").font = subtitle_font
    current_row += 2
    
    df_strat_dns = df_strategies[df_strategies['dataset'] == 'dns'][['Strategy', 'XGB threshold', 'IF threshold', 'Attack ratio', 'Confidence', 'Selected']]
    apply_styled_table(
        ws4, current_row, df_strat_dns,
        ["Strategy", "XGB threshold", "IF threshold", "Attack ratio", "Confidence score", "Selected strategy"],
        [align_left, align_right, align_right, align_right, align_right, align_center],
        [None, "0.00", "0.00", "0.0000", "0.0000", None]
    )
    current_row += len(df_strat_dns) + 2
    
    # Table B: Rescale vs Retrain Comparison (DNS)
    ws4.cell(row=current_row, column=1, value="Table B: Rescale vs Retrain Comparison (DNS)").font = table_title_font
    ws4.cell(row=current_row + 1, column=1, value="Evaluated on the next 40,000 test flows.").font = subtitle_font
    current_row += 2
    
    rows_b_dns = []
    for idx, r in df_results[df_results['dataset'] == 'dns'].iterrows():
        imp = "-" if r['Mode'] == 'Baseline' else f"{r['Improvement']:+.4f}"
        rows_b_dns.append([
            f"{r['Model']} ({r['Task']})",
            r['Mode'],
            r['Accuracy'],
            r['F1'],
            r['Recall'],
            r['FPR'],
            imp,
            f"{r['Cost']:.2f}s"
        ])
    df_table_b_dns = pd.DataFrame(rows_b_dns, columns=["Model (Task)", "Mode", "Accuracy", "F1", "Recall", "FPR", "Absolute F1 Improvement", "Training cost"])
    apply_styled_table(
        ws4, current_row, df_table_b_dns,
        df_table_b_dns.columns.tolist(),
        [align_left, align_left, align_right, align_right, align_right, align_right, align_right, align_right],
        [None, None, "0.000000", "0.000000", "0.000000", "0.000000", None, None]
    )
    current_row += len(df_table_b_dns) + 3
    
    # ----------------------------------------------------
    # Dataset: FRIDAY
    # ----------------------------------------------------
    ws4.cell(row=current_row, column=1, value="Dataset: FRIDAY").font = section_font
    current_row += 2
    
    # Table A: Calibration Strategy Selection (FRIDAY)
    ws4.cell(row=current_row, column=1, value="Table A: Calibration Strategy Selection (FRIDAY)").font = table_title_font
    ws4.cell(row=current_row + 1, column=1, value="Evaluated on the first 20,000 flows. Objective: avoid extreme ratio collapse.").font = subtitle_font
    current_row += 2
    
    df_strat_friday = df_strategies[df_strategies['dataset'] == 'friday'][['Strategy', 'XGB threshold', 'IF threshold', 'Attack ratio', 'Confidence', 'Selected']]
    apply_styled_table(
        ws4, current_row, df_strat_friday,
        ["Strategy", "XGB threshold", "IF threshold", "Attack ratio", "Confidence score", "Selected strategy"],
        [align_left, align_right, align_right, align_right, align_right, align_center],
        [None, "0.00", "0.00", "0.0000", "0.0000", None]
    )
    current_row += len(df_strat_friday) + 2
    
    # Table B: Rescale vs Retrain Comparison (FRIDAY)
    ws4.cell(row=current_row, column=1, value="Table B: Rescale vs Retrain Comparison (FRIDAY)").font = table_title_font
    ws4.cell(row=current_row + 1, column=1, value="Evaluated on the next 40,000 test flows.").font = subtitle_font
    current_row += 2
    
    rows_b_friday = []
    for idx, r in df_results[df_results['dataset'] == 'friday'].iterrows():
        imp = "-" if r['Mode'] == 'Baseline' else f"{r['Improvement']:+.4f}"
        rows_b_friday.append([
            f"{r['Model']} ({r['Task']})",
            r['Mode'],
            r['Accuracy'],
            r['F1'],
            r['Recall'],
            r['FPR'],
            imp,
            f"{r['Cost']:.2f}s"
        ])
    df_table_b_friday = pd.DataFrame(rows_b_friday, columns=["Model (Task)", "Mode", "Accuracy", "F1", "Recall", "FPR", "Absolute F1 Improvement", "Training cost"])
    apply_styled_table(
        ws4, current_row, df_table_b_friday,
        df_table_b_friday.columns.tolist(),
        [align_left, align_left, align_right, align_right, align_right, align_right, align_right, align_right],
        [None, None, "0.000000", "0.000000", "0.000000", "0.000000", None, None]
    )
    current_row += len(df_table_b_friday) + 3
    
    # ----------------------------------------------------
    # Table C: Key Insight Summary
    # ----------------------------------------------------
    ws4.cell(row=current_row, column=1, value="Table C: Key Insight Summary (Computed from Data)").font = table_title_font
    current_row += 1
    
    table_c_data = [
        ["Which dataset benefits more from rescale?", better_rescale_ds],
        ["Which model collapses without adaptation?", collapsing_model],
        ["Does retraining consistently outperform rescale?", outperform_str],
        ["When is rescale sufficient?", rescale_sufficient]
    ]
    df_table_c = pd.DataFrame(table_c_data, columns=["Question", "Computed Insight"])
    apply_styled_table(
        ws4, current_row, df_table_c,
        df_table_c.columns.tolist(),
        [align_left, align_left],
        [None, None]
    )
    
    autofit_columns(ws4)

    # Sheet 5: Controller
    print("[*] Generating Sheet 5: Controller Performance...")
    ws5 = wb.create_sheet(title="Controller")
    ws5.views.sheetView[0].showGridLines = True
    ws5.cell(row=1, column=1, value="SDN Controller Architecture Evaluation Summary").font = title_font
    ws5.cell(row=2, column=1, value="Comparative performance across L2 learning switches and AI-driven controllers (NRS, SPS, QPS, OFS)").font = subtitle_font
    
    # 1) Table 1: Switch vs C1
    ws5.cell(row=4, column=1, value="Table 1: Small Topology Performance Comparison (Simple L2 Switch vs. C1 Reactive AI)").font = bold_font
    c1_data = [
        ["probe", 0.9396, 0.6255, 0.6500, 0.8775],
        ["dos", 0.7001, 0.5971, 0.6500, 0.8775],
        ["ddos", 0.7704, 0.6255, 0.6500, 0.8775],
        ["sqli_web", 0.9396, 0.5493, 0.6500, 0.8775],
        ["credential_attack", 0.9396, 0.5486, 0.6500, 0.8775],
        ["exfiltration", 0.9396, 0.5504, 0.4875, 0.8775]
    ]
    c1_cols = ["Scenario", "Simple L2 Switch NRS", "Reactive AI (C1) NRS", "Simple L2 Switch SPS", "Reactive AI (C1) SPS"]
    c1_df = pd.DataFrame(c1_data, columns=c1_cols)
    for col_idx, col_name in enumerate(c1_df.columns):
        ws5.cell(row=5, column=col_idx + 1, value=col_name)
    apply_table_styles(ws5, 5, c1_df)
    
    # 2) Table 2: C2 vs C4 QPS/NRS
    start_row = 5 + len(c1_df) + 3
    ws5.cell(row=start_row, column=1, value="Table 2: Scale Drift Comparison: Controller 2 (Static) vs. Controller 4 (Feedback)").font = bold_font
    c2_c4_data = [
        ["Controller 2 - probe", 1.0000, 0.8800, 0.6230, 0.5240],
        ["Controller 2 - dos", 1.0000, 0.5600, 0.5911, 0.3785],
        ["Controller 2 - ddos", 1.0000, 0.8067, 0.6350, 0.5099],
        ["Controller 2 - sqli_web", 1.0000, 0.8467, 0.5494, 0.4698],
        ["Controller 2 - credential_attack", 1.0000, 0.6267, 0.5509, 0.3790],
        ["Controller 2 - exfiltration", 1.0000, 0.6667, 0.5476, 0.3983],
        ["Controller 4 - probe", 0.9400, 0.4933, 0.5784, 0.3881],
        ["Controller 4 - dos", 0.8000, 0.5333, 0.5845, 0.3318],
        ["Controller 4 - ddos", 0.8133, 0.4800, 0.5541, 0.4130],
        ["Controller 4 - sqli_web", 0.0333, 0.4733, 0.2900, 0.3147],
        ["Controller 4 - credential_attack", 0.0333, 0.4600, 0.2913, 0.3301],
        ["Controller 4 - exfiltration", 0.0333, 0.4933, 0.2923, 0.3351],
    ]
    c2_c4_cols = ["Controller & Scenario", "QPS (Small Topology)", "QPS (Large Topology)", "NRS (Small Topology)", "NRS (Large Topology)"]
    c2_c4_df = pd.DataFrame(c2_c4_data, columns=c2_c4_cols)
    for col_idx, col_name in enumerate(c2_c4_df.columns):
        ws5.cell(row=start_row + 1, column=col_idx + 1, value=col_name)
    apply_table_styles(ws5, start_row + 1, c2_c4_df)
    
    # 3) Table 3: C2 vs C3 Stability
    start_row_2 = start_row + 1 + len(c2_c4_df) + 3
    ws5.cell(row=start_row_2, column=1, value="Table 3: Topology Stability: C2 (Static Path) vs. C3 (Adaptive GNN Map)").font = bold_font
    c3_data = [
        ["C2 (XGB+IF)", 0.5828, 0.4433, -0.1395, 0.8775, 0.6780],
        ["C3 (XGB+IF+GNN)", 0.5425, 0.4705, -0.0720, 0.6500, 0.7083]
    ]
    c3_cols = ["Controller", "Mean NRS (Small)", "Mean NRS (Large)", "NRS Delta (Large - Small)", "Mean SPS (Small)", "Mean SPS (Large)"]
    c3_df = pd.DataFrame(c3_data, columns=c3_cols)
    for col_idx, col_name in enumerate(c3_df.columns):
        ws5.cell(row=start_row_2 + 1, column=col_idx + 1, value=col_name)
    apply_table_styles(ws5, start_row_2 + 1, c3_df)

    # 4) Table 4: C4 Small Scenario details
    start_row_3 = start_row_2 + 1 + len(c3_df) + 3
    ws5.cell(row=start_row_3, column=1, value="Table 4: Controller 4 Detailed Performance across Scenarios (Small Topology)").font = bold_font
    c4_data = [
        ["probe", 0.9400, 0.5784, 0.8775, 0.6500, 1.0000],
        ["dos", 0.8000, 0.5845, 0.8775, 0.6500, 1.0000],
        ["ddos", 0.8133, 0.5541, 0.8775, 0.6500, 1.0000],
        ["sqli_web", 0.0333, 0.2899, 0.8425, 0.5500, 1.0000],
        ["credential_attack", 0.0333, 0.2913, 0.7900, 0.4000, 1.0000],
        ["exfiltration", 0.0333, 0.2923, 0.7150, 0.6500, 0.7500]
    ]
    c4_cols = ["Scenario", "QPS Mean", "NRS Mean", "SPS Mean", "WS Mean", "DB Mean"]
    c4_df = pd.DataFrame(c4_data, columns=c4_cols)
    for col_idx, col_name in enumerate(c4_df.columns):
        ws5.cell(row=start_row_3 + 1, column=col_idx + 1, value=col_name)
    apply_table_styles(ws5, start_row_3 + 1, c4_df)

    # 5) Table 5: Raw Controller benchmark statistics (aggregated summary.csv)
    start_row_4 = start_row_3 + 1 + len(c4_df) + 3
    ws5.cell(row=start_row_4, column=1, value="Table 5: Raw Controller Benchmark Runs (Summary statistic aggregated over seeds)").font = bold_font
    summary_df = pd.read_csv('backend/benchmark/results/summary.csv')
    for col_idx, col_name in enumerate(summary_df.columns):
        ws5.cell(row=start_row_4 + 1, column=col_idx + 1, value=col_name)
    apply_table_styles(ws5, start_row_4 + 1, summary_df)
    autofit_columns(ws5)

    # Sheet 6: Features_order
    print("[*] Generating Sheet 6: Features_order...")
    ws6 = wb.create_sheet(title="Features_order")
    ws6.views.sheetView[0].showGridLines = True
    ws6.cell(row=1, column=1, value="GNN Features — Reference Orders").font = title_font
    ws6.cell(row=2, column=1, value="The lists of features in the exact column sequence expected during training").font = subtitle_font
    
    # Compile the lists of 15 features and 51 features
    # Fill empty values so they can be turned into a dataframe
    len_diff = len(feature_cols_51) - len(feature_cols_15)
    feature_cols_15_extended = feature_cols_15 + [''] * len_diff
    features_df = pd.DataFrame({
        'Selected 15 Features (Exact Order)': feature_cols_15_extended,
        'Selected 51 Features (Exact Order)': feature_cols_51
    })
    
    for col_idx, col_name in enumerate(features_df.columns):
        ws6.cell(row=4, column=col_idx + 1, value=col_name)
    apply_table_styles(ws6, 4, features_df)
    autofit_columns(ws6)

    # Sheet 7: Key Info & Scaler Analysis
    print("[*] Generating Sheet 7: Key Info & Scaler Analysis...")
    ws7 = wb.create_sheet(title="Key Info & Scaler Analysis")
    ws7.views.sheetView[0].showGridLines = True
    ws7.cell(row=1, column=1, value="Feature Scalers & Tri-Channel Calibration Analysis").font = title_font
    ws7.cell(row=2, column=1, value="Explores the design, contribution, and comparative analysis of Standard, Robust, and Tri-Channel scalers").font = subtitle_font
    
    ws7.cell(row=4, column=1, value="1. GNN Model Performance Comparison: Standard Scaler vs. Robust Scaler (Mean & Max Test F1)").font = bold_font
    scaler_comp_data = [
        ["Binary", "GAT", 0.996989, 0.998402, 0.897931, 0.995072],
        ["Binary", "GraphSAGE", 0.996383, 0.998275, 0.950791, 0.994083],
        ["Multiclass", "GAT", 0.448478, 0.486118, 0.247969, 0.348685],
        ["Multiclass", "GraphSAGE", 0.510776, 0.598308, 0.291421, 0.467994]
    ]
    scaler_comp_cols = ["Task Type", "Architecture", "Standard Scaler F1 (Mean)", "Standard Scaler F1 (Max)", "Robust Scaler F1 (Mean)", "Robust Scaler F1 (Max)"]
    scaler_comp_df = pd.DataFrame(scaler_comp_data, columns=scaler_comp_cols)
    for col_idx, col_name in enumerate(scaler_comp_df.columns):
        ws7.cell(row=5, column=col_idx + 1, value=col_name)
    apply_table_styles(ws7, 5, scaler_comp_df)
    
    txt_start = 5 + len(scaler_comp_df) + 3
    ws7.cell(row=txt_start, column=1, value="2. Scaler Concept & Contribution Analysis").font = bold_font
    
    commentaries = [
        ("What is the Tri-Channel Scaler and how does it contribute?", 
         "The Tri-Channel Scaler transforms each continuous network traffic feature into three distinct representation channels:\n"
         "  1) Log Channel [log1p(x)]: Captures the non-linear distribution shape of features. This helps compress highly skewed network volume metrics (e.g. Flow Duration, Packet Lengths) to stabilize learning.\n"
         "  2) Ratio Channel [x / P95_benign]: Provides scale invariance. By dividing by the 95th percentile of benign traffic, it maps features relative to a benign baseline rate. This ensures the features generalize across topologies of different sizes, where raw traffic rates naturally shift.\n"
         "  3) Delta Channel [(x - Median) / IQR_benign]: Normalizes features based on robust statistics to highlight standard deviations. This creates a clear signal representing how far a flow deviates from normal, which is highly effective for anomaly detectors (like Isolation Forest) to separate attacks from benign flows."),
        
        ("Is the contribution limited to only rescaling?", 
         "No, it does far more than simple rescaling:\n"
         "  * Logarithmic dampening alters the non-linear relationship of features, preventing extreme outliers from skewing model weights.\n"
         "  * Feature expansion (from 15 features to 45 features) structures distinct characteristics (shape, ratio, deviation) explicitly. This allows classifiers (XGBoost, Isolation Forest) to learn robust 'environment-agnostic' patterns rather than rigid absolute values.\n"
         "  * Most crucially, it acts as an environment calibration boundary. When deployed on a new topology, the baseline stats (benign median, IQR, P95) are dynamically re-calculated on local benign traffic. This rescales the live traffic dynamically to match the training environment's representation, neutralizing Scale Drift."),
        
        ("Can Robust and Standard Scalers rescale also help?", 
         "Yes, they help models converge, but they have major limitations:\n"
         "  * Standard Scaler: Standardizes features using Mean and Standard Deviation. Since network traffic features are highly skewed and contain massive outliers (especially under flooding attacks), the Mean and StdDev are heavily distorted. Standard Scaler fails to preserve clean classification boundaries under outliers.\n"
         "  * Robust Scaler: Uses Median and IQR, which makes it far more robust to outliers because it relies on percentiles. It is superior to Standard Scaler for cleaning skewed data.\n"
         "  * Limitations: Standard and Robust scalers only center and scale features. They do not generate multi-channel representations (log and ratio), and they do not provide scale invariance when transitioning to larger topologies (scale shift) unless dynamically refitted on the controller.\n"
         "  * GNN Evidence: In standard GNN evaluations, GNN Standard Scaler (mean F1 ~0.51 multiclass, max F1 ~0.59) outperformed GNN Robust Scaler (mean F1 ~0.29 multiclass, max F1 ~0.46) on the standard dataset, but both suffered from scale drift under topological changes. Standard scaled models collapsed (-0.1395 NRS degradation), while models adapting the Tri-Channel scaler locally maintained topological stability (-0.0720 NRS delta, a 50% improvement).")
    ]
    
    curr_row = txt_start + 1
    for title, content in commentaries:
        ws7.cell(row=curr_row, column=1, value=title).font = bold_font
        ws7.row_dimensions[curr_row].height = 20
        curr_row += 1
        
        ws7.cell(row=curr_row, column=1, value=content).font = normal_font
        ws7.cell(row=curr_row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        ws7.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row+6, end_column=6)
        ws7.row_dimensions[curr_row].height = 120
        curr_row += 8
        
    autofit_columns(ws7)
    ws7.column_dimensions['A'].width = 30
    ws7.column_dimensions['B'].width = 18
    ws7.column_dimensions['C'].width = 25
    ws7.column_dimensions['D'].width = 25
    ws7.column_dimensions['E'].width = 25
    ws7.column_dimensions['F'].width = 25

    output_path = "Evaluation_Results_Compilation.xlsx"
    wb.save(output_path)
    print(f"[+] Workbook successfully saved to {output_path}")

if __name__ == "__main__":
    main()
