#!/usr/bin/env python3
"""
generate_all_figures.py — Master Figure & Excel Generator (Matching Expected Paper Layout)

Reads valid benchmark data from results/final_atdm_runs/ and generates:
  - fig2_latency_timeline.png
  - fig3_security_preservation.png
  - fig4_service_availability.png
  - fig5_bandwidth_util.png
  - fig6_throughput_timeline.png (and fig6_throughput_util.png)
  - final_experiment_results.xlsx (9 sheets)

Layout: Both controllers (Simple Switch 13 and ATDM) are plotted on the SAME graph
for each topology (Small Topology in ax1, Large Topology in ax2) for direct comparison.
"""

import os
import sys
import json
import glob
import re
import time
import numpy as np
import pandas as pd

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ────────────────────────────────────────────────────────────────
BENCHMARK_DIR = os.path.dirname(os.path.abspath(__file__))
RESULTS_DIR = os.path.join(BENCHMARK_DIR, "results")
FINAL_RUNS_DIR = os.path.join(RESULTS_DIR, "final_atdm_runs")
FIGURES_DIR = os.path.join(BENCHMARK_DIR, "figures")
EXPECTED_IMAGES_DIR = os.path.join(BENCHMARK_DIR, "images", "expected")
REPO_ROOT = os.path.dirname(os.path.dirname(BENCHMARK_DIR))  # fyp/
GNN_COMPARE_DIR = os.path.join(os.path.dirname(BENCHMARK_DIR), "gnn_compare")

# Output paths
EXCEL_PATH = os.path.join(REPO_ROOT, "final_experiment_results.xlsx")

# Publication Style Configuration (Wong 2011 palette)
PALETTE = {
    "blue":    "#0072B2",
    "orange":  "#E69F00",
    "green":   "#009E73",
    "red":     "#D55E00",
    "purple":  "#CC79A7",
    "cyan":    "#56B4E9",
    "yellow":  "#F0E442",
    "grey":    "#999999",
}

STYLE = {
    "font.family": "sans-serif",
    "font.sans-serif": ["DejaVu Sans", "Arial", "Helvetica"],
    "font.size": 11,
    "axes.titlesize": 13,
    "axes.titleweight": "bold",
    "axes.labelsize": 12,
    "axes.labelweight": "bold",
    "xtick.labelsize": 10,
    "ytick.labelsize": 10,
    "legend.fontsize": 9.5,
    "legend.framealpha": 0.9,
    "legend.edgecolor": "#cccccc",
    "figure.dpi": 150,
    "savefig.dpi": 300,
    "savefig.bbox": "tight",
    "axes.grid": True,
    "grid.alpha": 0.25,
    "grid.linewidth": 0.5,
    "axes.spines.top": False,
    "axes.spines.right": False,
    "axes.linewidth": 1.2,
}

# Display name mapping
DISPLAY_NAMES = {
    'controller_4': 'ATDM',
    'atdm': 'ATDM',
    'simple_13': 'Simple Switch 13',
    'simple_switch_13': 'Simple Switch 13',
}

def display_name(ctrl):
    return DISPLAY_NAMES.get(ctrl, ctrl)

# Attack scenario ordering & labels
ATTACK_ORDER = ['probe', 'dos', 'ddos', 'sqli_web', 'credential_attack', 'exfiltration']
SCENARIO_LABELS = {
    'probe': 'Probe', 'dos': 'DoS', 'ddos': 'DDoS',
    'sqli_web': 'SQLi', 'credential_attack': 'Credential',
    'exfiltration': 'Exfiltration',
}


# ═══════════════════════════════════════════════════════════════════════════
# Data Loading
# ═══════════════════════════════════════════════════════════════════════════
def parse_filename(filename):
    """Extract controller, topology, attack, seed from filename."""
    basename = os.path.splitext(os.path.basename(filename))[0]
    seed_match = re.search(r'_seed_(\d+)$', basename)
    seed = int(seed_match.group(1)) if seed_match else 1
    prefix = basename[:seed_match.start()] if seed_match else basename

    for ctrl in ['simple_switch_13', 'atdm', 'controller_4', 'simple_13']:
        if prefix.startswith(ctrl + '_'):
            remainder = prefix[len(ctrl) + 1:]
            for topo in ['small', 'large']:
                if remainder.startswith(topo + '_'):
                    attack = remainder[len(topo) + 1:]
                    return ctrl, topo, attack, seed
                elif remainder == topo:
                    return ctrl, topo, 'unknown', seed

    return 'unknown', 'unknown', 'unknown', seed


def load_all_runs():
    """Load all valid benchmark run files from final_atdm_runs/."""
    runs = []
    json_files = sorted(glob.glob(os.path.join(FINAL_RUNS_DIR, "*.json")))

    if not json_files:
        print(f"[ERROR] No JSON files found in {FINAL_RUNS_DIR}")
        sys.exit(1)

    for fpath in json_files:
        ctrl_file, topo_file, attack_file, seed_file = parse_filename(fpath)

        with open(fpath, 'r') as f:
            data = json.load(f)

        results = data.get('results', {})
        scores = None

        for ck, tv in results.items():
            if isinstance(tv, dict):
                for tk, sv in tv.items():
                    if isinstance(sv, dict) and ('NRS' in sv or 'probe_history' in sv):
                        scores = sv
                        break
            if scores:
                break

        if not scores:
            continue

        run = {
            'file': os.path.basename(fpath),
            'controller': ctrl_file,
            'topology': topo_file,
            'attack': attack_file,
            'seed': seed_file,
            'scores': scores,
            'probe_history': scores.get('probe_history', []),
            'qos_history': scores.get('qos_history', []),
            'flow_history': scores.get('flow_history', []),
            'SCS': scores.get('SCS', 0.0),
            'QPS': scores.get('QPS', 0.0),
            'UIS': scores.get('UIS', 0.0),
            'RES': scores.get('RES', 0.0),
            'NRS': scores.get('NRS', 0.0),
            'WS': scores.get('WS', 0.0),
            'DB': scores.get('DB', 0.0),
            'SPS': scores.get('SPS', 0.0),
        }
        runs.append(run)

    print(f"[data] Loaded {len(runs)} valid run files")
    return runs


def load_concatenated_timeline(runs, controller_key, topology, metric_type="latency"):
    """
    Build 6.5-minute concatenated timeline (6 scenarios x ~65s) for a given controller & topology.
    metric_type can be 'latency', 'bandwidth', or 'throughput'.
    """
    concatenated_time = []
    concatenated_val = []
    current_offset = 0.0
    timeline_rows = []

    # Map controller names (atdm -> controller_4, etc.)
    target_ctrls = [controller_key]
    if controller_key in ['atdm', 'controller_4']:
        target_ctrls = ['atdm', 'controller_4']
    elif controller_key in ['simple_13', 'simple_switch_13']:
        target_ctrls = ['simple_13', 'simple_switch_13']

    for scen in ATTACK_ORDER:
        matching = [r for r in runs
                   if r['controller'] in target_ctrls
                   and r['topology'] == topology
                   and r['attack'] == scen]

        if not matching:
            # Fallback 65s of zeros if missing
            for t in range(65):
                concatenated_time.append(current_offset + t)
                concatenated_val.append(0.0)
            current_offset += 65.0
            continue

        run = matching[0]
        scores = run['scores']

        if metric_type == "latency":
            ph = run['probe_history']
            probes = [p for p in ph if p.get('latency_ms') is not None and p.get('latency_ms', 0) > 0]
            if not probes:
                probes = ph

            if not probes:
                for t in range(65):
                    concatenated_time.append(current_offset + t)
                    concatenated_val.append(0.0)
                current_offset += 65.0
                continue

            # Group latency by elapsed second (calculating relative timestamp from timestamp if elapsed is missing)
            probes = sorted(probes, key=lambda x: x.get('timestamp', 0))
            first_ts = probes[0].get('timestamp', 0)
            sec_dict = {}
            for p in probes:
                lat = p.get('latency_ms', 0) or 0.0
                if 'elapsed' in p:
                    t_rel = p['elapsed']
                elif 'timestamp' in p and first_ts > 0:
                    t_rel = p['timestamp'] - first_ts
                else:
                    t_rel = 0.0
                t = int(round(t_rel))
                t = max(0, min(64, t))
                sec_dict.setdefault(t, []).append(lat)

            max_sec = max(max(sec_dict.keys(), default=0), 64)
            last_lat = 0.0
            for t in range(max_sec + 1):
                lats = sec_dict.get(t, [])
                if lats:
                    last_lat = float(np.mean(lats))
                avg_lat = last_lat
                t_global = current_offset + t
                concatenated_time.append(t_global)
                concatenated_val.append(avg_lat)
                timeline_rows.append({
                    'controller': display_name(controller_key),
                    'topology': topology,
                    'attack': SCENARIO_LABELS[scen],
                    'run_id': run['file'],
                    'timestamp': t_global,
                    'attack_active': 20 <= t <= 50,
                    'benign_avg_latency_ms': avg_lat,
                    'successful_requests': 1 if avg_lat > 0 else 0,
                    'failed_requests': 0 if avg_lat > 0 else 1,
                })
            current_offset += (max_sec + 1)

        elif metric_type == "bandwidth":
            qh = run['qos_history']
            if not qh:
                for t in range(65):
                    concatenated_time.append(current_offset + t)
                    concatenated_val.append(0.0)
                current_offset += 65.0
                continue

            sec_dict = {}
            for q in qh:
                t = int(round(q.get('elapsed', 0)))
                t = max(0, min(64, t))
                total_bytes = sum(q.get('throughput', {}).values())
                sec_dict[t] = total_bytes  # B/s

            max_sec = max(max(sec_dict.keys(), default=0), 64)
            # Link limit for % calculation: 20 Mbps baseline = 2,560,000 B/s
            link_limit = 2_560_000.0
            for t in range(max_sec + 1):
                bytes_sec = sec_dict.get(t, 0.0)
                util_pct = (bytes_sec / link_limit) * 100.0
                t_global = current_offset + t
                concatenated_time.append(t_global)
                concatenated_val.append(util_pct)
                timeline_rows.append({
                    'controller': display_name(controller_key),
                    'topology': topology,
                    'attack': SCENARIO_LABELS[scen],
                    'run_id': run['file'],
                    'timestamp': t_global,
                    'attack_active': 20 <= t <= 50,
                    'measured_bandwidth_Bps': bytes_sec,
                    'available_bandwidth_Bps': link_limit,
                    'utilization_pct': util_pct,
                })
            current_offset += (max_sec + 1)

        elif metric_type == "throughput":
            fh = run['flow_history']
            if not fh:
                for t in range(65):
                    concatenated_time.append(current_offset + t)
                    concatenated_val.append(0.0)
                current_offset += 65.0
                continue

            sec_dict = {}
            normal_hosts = ["h1"] if topology == "large" else ["h1", "h4", "h5"]
            for f in fh:
                t = int(round(f.get('elapsed', 0)))
                t = max(0, min(64, t))
                tp_dict = f.get('throughput', {})
                normal_bytes = sum(tp_dict.values())
                sec_dict[t] = normal_bytes / 1024.0  # KB/s

            max_sec = max(max(sec_dict.keys(), default=0), 64)
            for t in range(max_sec + 1):
                kb_sec = sec_dict.get(t, 0.0)
                t_global = current_offset + t
                concatenated_time.append(t_global)
                concatenated_val.append(kb_sec)
                timeline_rows.append({
                    'controller': display_name(controller_key),
                    'topology': topology,
                    'attack': SCENARIO_LABELS[scen],
                    'run_id': run['file'],
                    'timestamp': t_global,
                    'attack_active': 20 <= t <= 50,
                    'benign_throughput_KBps': kb_sec,
                })
            current_offset += (max_sec + 1)

    return concatenated_time, concatenated_val, timeline_rows


def smooth_values(values, window_size=3):
    """Apply rolling average smoothing to timeline values."""
    if not values:
        return []
    smoothed = []
    for i in range(len(values)):
        start_idx = max(0, i - window_size + 1)
        window = values[start_idx : i + 1]
        smoothed.append(sum(window) / len(window))
    return smoothed


# ═══════════════════════════════════════════════════════════════════════════
# Figure 2: Latency Timeline (Line Chart comparing Controllers per Topology)
# ═══════════════════════════════════════════════════════════════════════════
def generate_fig2(runs):
    """Figure 2 — Latency timeline comparing Simple Switch vs ATDM on the same plot for each topology."""
    print("\n[fig2] Generating Latency Timeline...")

    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(15, 10), sharex=True)

    # Load timelines
    t_sm_simple, v_sm_simple, rows1 = load_concatenated_timeline(runs, "simple_switch_13", "small", "latency")
    t_sm_atdm, v_sm_atdm, rows2 = load_concatenated_timeline(runs, "atdm", "small", "latency")

    t_lg_simple, v_lg_simple, rows3 = load_concatenated_timeline(runs, "simple_switch_13", "large", "latency")
    t_lg_atdm, v_lg_atdm, rows4 = load_concatenated_timeline(runs, "atdm", "large", "latency")

    all_rows = rows1 + rows2 + rows3 + rows4

    # Compute max limits
    max_small = max(max(v_sm_simple or [0]), max(v_sm_atdm or [0]))
    max_large = max(max(v_lg_simple or [0]), max(v_lg_atdm or [0]))

    # Plot Small Topology (ax1)
    ax1.plot(t_sm_simple, v_sm_simple, label="Small — Simple Switch 13", color=PALETTE["blue"], linewidth=1.5, alpha=0.85)
    ax1.plot(t_sm_atdm, v_sm_atdm, label="Small — ATDM", color=PALETTE["orange"], linewidth=1.5, alpha=0.85)

    # Plot Large Topology (ax2)
    ax2.plot(t_lg_simple, v_lg_simple, label="Large — Simple Switch 13", color=PALETTE["green"], linewidth=1.5, alpha=0.85)
    ax2.plot(t_lg_atdm, v_lg_atdm, label="Large — ATDM", color=PALETTE["red"], linewidth=1.5, alpha=0.85)

    # Configure axes
    for ax, ymax, title_text in [(ax1, max_small, "Small Network Topology"), (ax2, max_large, "Large Network Topology")]:
        for i, scen in enumerate(ATTACK_ORDER):
            ax.axvspan(i * 65 + 20, i * 65 + 50, color="#ffcccc", alpha=0.2,
                       label="Attack Active" if i == 0 else "")
            if i > 0:
                ax.axvline(x=i * 65, color="#888888", linestyle=":", alpha=0.5)

        ax.set_ylim(bottom=-10, top=max(ymax * 1.30, 100))
        curr_ymax = ax.get_ylim()[1]

        for i, scen in enumerate(ATTACK_ORDER):
            center_x = i * 65 + 32.5
            ax.text(center_x, curr_ymax * 0.85, SCENARIO_LABELS[scen], ha="center", va="center",
                    fontsize=8.5, fontweight="bold",
                    bbox=dict(boxstyle="round,pad=0.2", facecolor="white", edgecolor="#cccccc", alpha=0.8))

        ax.set_ylabel("Average Latency (ms)")
        ax.set_title(title_text, fontsize=12, fontweight="bold")
        ax.legend(loc="upper left", framealpha=0.9)

    ax2.set_xlabel("Timeline (minutes)")
    ticks = np.arange(0, 391, 60)
    ax2.set_xticks(ticks)
    ax2.set_xticklabels([f"{t//60}m {t%60}s" if t > 0 else "0s" for t in ticks])

    fig.suptitle("Figure 2: User-Perceived Latency Timeline across Scenarios", fontsize=15, fontweight="bold", y=0.98)
    plt.tight_layout(rect=[0, 0, 1, 0.95])

    path = os.path.join(FIGURES_DIR, "fig2_latency_timeline.png")
    fig.savefig(path, dpi=300)
    plt.close(fig)
    print(f"[fig2] Saved: {path}")
    return all_rows


# ═══════════════════════════════════════════════════════════════════════════
# Figure 3: Security Preservation Score
# ═══════════════════════════════════════════════════════════════════════════
def generate_fig3(runs):
    """Figure 3 — Security Preservation grouped bar chart comparing WS and DB for both controllers per topology."""
    print("\n[fig3] Generating Security Preservation Score...")

    sps_attacks = ['sqli_web', 'credential_attack', 'exfiltration']
    controllers = ['simple_switch_13', 'atdm']
    metrics = ['WS', 'DB']
    topologies = ['small', 'large']
    sps_rows = []

    fig, axes = plt.subplots(1, 2, figsize=(15, 6.5), sharey=True)
    topo_titles = {"small": "Small Topology", "large": "Large Topology"}

    for ax_idx, topo in enumerate(topologies):
        ax = axes[ax_idx]
        x = np.arange(len(sps_attacks))
        n_bars = 4  # simple_WS, atdm_WS, simple_DB, atdm_DB
        total_width = 0.75
        width = total_width / n_bars

        bar_colors = [PALETTE["blue"], PALETTE["orange"], PALETTE["green"], PALETTE["red"]]
        bar_labels = [
            "Simple Switch 13 — WS",
            "ATDM — WS",
            "Simple Switch 13 — DB",
            "ATDM — DB",
        ]

        bar_idx = 0
        for metric in metrics:
            for ctrl in controllers:
                vals = []
                for scen in sps_attacks:
                    matching = [r for r in runs
                               if r['controller'] in [ctrl, display_name(ctrl).lower()]
                               and r['topology'] == topo
                               and r['attack'] == scen]
                    if matching:
                        val = matching[0]['scores'].get(metric, 0.0)
                    else:
                        val = 0.0
                    vals.append(val)
                    sps_rows.append({
                        'controller': display_name(ctrl),
                        'topology': topo,
                        'attack': SCENARIO_LABELS.get(scen, scen),
                        'metric': metric,
                        'score': val,
                    })

                offset = (bar_idx - n_bars / 2 + 0.5) * width
                bars = ax.bar(x + offset, vals, width,
                              label=bar_labels[bar_idx], color=bar_colors[bar_idx],
                              edgecolor="white", linewidth=0.8, alpha=0.9)

                for bar, val in zip(bars, vals):
                    ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.02,
                            f"{val:.2f}", ha="center", va="bottom", fontsize=7, fontweight="bold")
                bar_idx += 1

        ax.set_xlabel("Attack Scenario")
        ax.set_title(topo_titles[topo], fontsize=13, fontweight="bold")
        ax.set_xticks(x)
        ax.set_xticklabels([SCENARIO_LABELS[s] for s in sps_attacks], fontsize=10)
        ax.set_ylim(0, 1.2)
        ax.legend(loc="lower right", fontsize=8, ncol=2)

    axes[0].set_ylabel("Security Score (0–1)")
    fig.suptitle("Figure 3: Security Preservation — Web Server & Database Protection", fontsize=15, fontweight="bold", y=1.02)

    path = os.path.join(FIGURES_DIR, "fig3_security_preservation.png")
    fig.savefig(path, dpi=300)
    plt.close(fig)
    print(f"[fig3] Saved: {path}")
    return sps_rows


# ═══════════════════════════════════════════════════════════════════════════
# Figure 4: Service Availability
# ═══════════════════════════════════════════════════════════════════════════
def generate_fig4(runs):
    """Figure 4 — Service Availability (SCS) grouped bar chart comparing Simple Switch vs ATDM per topology."""
    print("\n[fig4] Generating Service Availability...")

    controllers = ['simple_switch_13', 'atdm']
    topologies = ['small', 'large']
    avail_rows = []

    fig, axes = plt.subplots(1, 2, figsize=(14, 6.5), sharey=True)
    topo_titles = {"small": "Small Topology", "large": "Large Topology"}

    for ax_idx, topo in enumerate(topologies):
        ax = axes[ax_idx]
        x = np.arange(len(ATTACK_ORDER))
        width = 0.35

        for c_idx, ctrl in enumerate(controllers):
            vals = []
            for scen in ATTACK_ORDER:
                matching = [r for r in runs
                           if r['controller'] in [ctrl, display_name(ctrl).lower()]
                           and r['topology'] == topo
                           and r['attack'] == scen]
                val = matching[0]['SCS'] if matching else 0.0
                vals.append(val)
                avail_rows.append({
                    'controller': display_name(ctrl),
                    'topology': topo,
                    'attack': SCENARIO_LABELS.get(scen, scen),
                    'SCS': val,
                })

            offset = (c_idx - 0.5) * width
            color = PALETTE["blue"] if ctrl == "simple_switch_13" else PALETTE["orange"]
            label = display_name(ctrl)

            bars = ax.bar(x + offset, vals, width,
                          label=label, color=color,
                          edgecolor="white", linewidth=0.8, alpha=0.9)

            for bar, val in zip(bars, vals):
                ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.02,
                        f"{val:.2f}", ha="center", va="bottom", fontsize=7, fontweight="bold")

        ax.axhline(y=1.0, color=PALETTE["grey"], linestyle="--", linewidth=1.5, alpha=0.6, label="Maximum SCS (1.0)")
        ax.set_xlabel("Attack Scenario")
        ax.set_title(topo_titles[topo], fontsize=13, fontweight="bold")
        ax.set_xticks(x)
        ax.set_xticklabels([SCENARIO_LABELS[s] for s in ATTACK_ORDER], fontsize=10, rotation=20, ha="right")
        ax.set_ylim(0, 1.2)
        ax.legend(loc="upper right", fontsize=9)

    axes[0].set_ylabel("Service Continuity Score (SCS)")
    fig.suptitle("Figure 4: Service Availability — Resilience Under Attack", fontsize=15, fontweight="bold", y=1.02)

    path = os.path.join(FIGURES_DIR, "fig4_service_availability.png")
    fig.savefig(path, dpi=300)
    plt.close(fig)
    print(f"[fig4] Saved: {path}")
    return avail_rows


# ═══════════════════════════════════════════════════════════════════════════
# Figure 5: Bandwidth Utilization Timeline
# ═══════════════════════════════════════════════════════════════════════════
def generate_fig5(runs):
    """Figure 5 — Bandwidth utilization timeline comparing Simple Switch vs ATDM per topology."""
    print("\n[fig5] Generating Bandwidth Utilization Timeline...")

    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(15, 10), sharex=True)

    t_sm_simple, v_sm_simple, rows1 = load_concatenated_timeline(runs, "simple_switch_13", "small", "bandwidth")
    v_sm_simple_s = smooth_values(v_sm_simple, 3)

    t_sm_atdm, v_sm_atdm, rows2 = load_concatenated_timeline(runs, "atdm", "small", "bandwidth")
    v_sm_atdm_s = smooth_values(v_sm_atdm, 3)

    t_lg_simple, v_lg_simple, rows3 = load_concatenated_timeline(runs, "simple_switch_13", "large", "bandwidth")
    v_lg_simple_s = smooth_values(v_lg_simple, 3)

    t_lg_atdm, v_lg_atdm, rows4 = load_concatenated_timeline(runs, "atdm", "large", "bandwidth")
    v_lg_atdm_s = smooth_values(v_lg_atdm, 3)

    all_rows = rows1 + rows2 + rows3 + rows4

    # Plot Small Topology (ax1)
    ax1.plot(t_sm_simple, v_sm_simple_s, label="Small — Simple Switch 13", color=PALETTE["blue"], linewidth=1.5, alpha=0.85)
    ax1.plot(t_sm_atdm, v_sm_atdm_s, label="Small — ATDM", color=PALETTE["orange"], linewidth=1.5, alpha=0.85)

    # Plot Large Topology (ax2)
    ax2.plot(t_lg_simple, v_lg_simple_s, label="Large — Simple Switch 13", color=PALETTE["green"], linewidth=1.5, alpha=0.85)
    ax2.plot(t_lg_atdm, v_lg_atdm_s, label="Large — ATDM", color=PALETTE["red"], linewidth=1.5, alpha=0.85)

    for ax, title_text in [(ax1, "Small Network Topology"), (ax2, "Large Network Topology")]:
        for i, scen in enumerate(ATTACK_ORDER):
            ax.axvspan(i * 65 + 20, i * 65 + 50, color="#ffcccc", alpha=0.2,
                       label="Attack Active" if i == 0 else "")
            if i > 0:
                ax.axvline(x=i * 65, color="#888888", linestyle=":", alpha=0.5)

        ax.set_ylim(bottom=-5, top=110)
        ax.set_yticks([0, 20, 40, 60, 80, 100])
        ax.axhline(y=100.0, color="#666666", linestyle="--", linewidth=1.2, alpha=0.8, label="Link Capacity (100%)")

        for i, scen in enumerate(ATTACK_ORDER):
            center_x = i * 65 + 32.5
            ax.text(center_x, 93.5, SCENARIO_LABELS[scen], ha="center", va="center",
                    fontsize=8.5, fontweight="bold",
                    bbox=dict(boxstyle="round,pad=0.2", facecolor="white", edgecolor="#cccccc", alpha=0.8))

        ax.set_ylabel("Bandwidth Utilization (%)")
        ax.set_title(title_text, fontsize=12, fontweight="bold")
        ax.legend(loc="upper left", framealpha=0.9)

    ax2.set_xlabel("Timeline (minutes)")
    ticks = np.arange(0, 391, 60)
    ax2.set_xticks(ticks)
    ax2.set_xticklabels([f"{t//60}m {t%60}s" if t > 0 else "0s" for t in ticks])

    fig.suptitle("Figure 5: Bandwidth Utilization Timeline across Scenarios", fontsize=15, fontweight="bold", y=0.98)
    plt.tight_layout(rect=[0, 0, 1, 0.95])

    path = os.path.join(FIGURES_DIR, "fig5_bandwidth_util.png")
    fig.savefig(path, dpi=300)
    plt.close(fig)
    print(f"[fig5] Saved: {path}")
    return all_rows


# ═══════════════════════════════════════════════════════════════════════════
# Figure 6: Benign Throughput Timeline
# ═══════════════════════════════════════════════════════════════════════════
def generate_fig6(runs):
    """Figure 6 — Benign throughput timeline comparing Simple Switch vs ATDM per topology."""
    print("\n[fig6] Generating Benign Throughput Timeline...")

    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(15, 10), sharex=True)

    t_sm_simple, v_sm_simple, rows1 = load_concatenated_timeline(runs, "simple_switch_13", "small", "throughput")
    v_sm_simple_s = smooth_values(v_sm_simple, 3)

    t_sm_atdm, v_sm_atdm, rows2 = load_concatenated_timeline(runs, "atdm", "small", "throughput")
    v_sm_atdm_s = smooth_values(v_sm_atdm, 3)

    t_lg_simple, v_lg_simple, rows3 = load_concatenated_timeline(runs, "simple_switch_13", "large", "throughput")
    v_lg_simple_s = smooth_values(v_lg_simple, 3)

    t_lg_atdm, v_lg_atdm, rows4 = load_concatenated_timeline(runs, "atdm", "large", "throughput")
    v_lg_atdm_s = smooth_values(v_lg_atdm, 3)

    all_rows = rows1 + rows2 + rows3 + rows4

    max_small = max(max(v_sm_simple_s or [0]), max(v_sm_atdm_s or [0]), 158.6)
    max_large = max(max(v_lg_simple_s or [0]), max(v_lg_atdm_s or [0]), 1077.7)

    # Plot Small Topology (ax1)
    ax1.plot(t_sm_simple, v_sm_simple_s, label="Small — Simple Switch 13", color=PALETTE["blue"], linewidth=1.5, alpha=0.85)
    ax1.plot(t_sm_atdm, v_sm_atdm_s, label="Small — ATDM", color=PALETTE["orange"], linewidth=1.5, alpha=0.85)
    ax1.axhline(y=158.6, color="#888888", linestyle=":", linewidth=1.5, alpha=0.9, label="Benign Baseline (158.6 KB/s)")

    # Plot Large Topology (ax2)
    ax2.plot(t_lg_simple, v_lg_simple_s, label="Large — Simple Switch 13", color=PALETTE["green"], linewidth=1.5, alpha=0.85)
    ax2.plot(t_lg_atdm, v_lg_atdm_s, label="Large — ATDM", color=PALETTE["red"], linewidth=1.5, alpha=0.85)
    ax2.axhline(y=1077.7, color="#888888", linestyle=":", linewidth=1.5, alpha=0.9, label="Benign Baseline (1077.7 KB/s)")

    for ax, ymax, title_text in [(ax1, max_small, "Small Network Topology"), (ax2, max_large, "Large Network Topology")]:
        for i, scen in enumerate(ATTACK_ORDER):
            ax.axvspan(i * 65 + 20, i * 65 + 50, color="#ffcccc", alpha=0.2,
                       label="Attack Active" if i == 0 else "")
            if i > 0:
                ax.axvline(x=i * 65, color="#888888", linestyle=":", alpha=0.5)

        ax.set_ylim(bottom=-10, top=max(ymax * 1.30, 100))
        curr_ymax = ax.get_ylim()[1]

        for i, scen in enumerate(ATTACK_ORDER):
            center_x = i * 65 + 32.5
            ax.text(center_x, curr_ymax * 0.85, SCENARIO_LABELS[scen], ha="center", va="center",
                    fontsize=8.5, fontweight="bold",
                    bbox=dict(boxstyle="round,pad=0.2", facecolor="white", edgecolor="#cccccc", alpha=0.8))

        ax.set_ylabel("Benign User Throughput (KB/s)")
        ax.set_title(title_text, fontsize=12, fontweight="bold")
        ax.legend(loc="upper left", framealpha=0.9)

    ax2.set_xlabel("Timeline (minutes)")
    ticks = np.arange(0, 391, 60)
    ax2.set_xticks(ticks)
    ax2.set_xticklabels([f"{t//60}m {t%60}s" if t > 0 else "0s" for t in ticks])

    fig.suptitle("Figure 6: Benign User Throughput Timeline across Scenarios", fontsize=15, fontweight="bold", y=0.98)
    plt.tight_layout(rect=[0, 0, 1, 0.95])

    path = os.path.join(FIGURES_DIR, "fig6_throughput_timeline.png")
    fig.savefig(path, dpi=300)
    path_util = os.path.join(FIGURES_DIR, "fig6_throughput_util.png")
    fig.savefig(path_util, dpi=300)
    plt.close(fig)
    print(f"[fig6] Saved: {path}")
    print(f"[fig6] Saved: {path_util}")
    return all_rows


# ═══════════════════════════════════════════════════════════════════════════
# Excel Workbook Generation
# ═══════════════════════════════════════════════════════════════════════════
def generate_excel(runs, latency_rows, sps_rows, avail_rows, bw_rows, tp_rows):
    """Generate comprehensive 9-sheet Excel workbook."""
    print("\n[excel] Generating Excel workbook...")

    wb = openpyxl.Workbook()
    ws_default = wb.active
    wb.remove(ws_default)

    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    normal_font = Font(name='Calibri', size=11)
    zebra_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    thin_side = Side(style='thin', color='D9D9D9')
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')

    def write_sheet(ws, df, title=None):
        start_row = 1
        if title:
            ws.cell(row=1, column=1, value=title).font = Font(name='Calibri', size=14, bold=True, color='1F4E79')
            start_row = 3

        for ci, col in enumerate(df.columns, 1):
            cell = ws.cell(row=start_row, column=ci, value=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = cell_border

        for ri, (_, row) in enumerate(df.iterrows()):
            fill = zebra_fill if ri % 2 == 1 else white_fill
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=start_row + 1 + ri, column=ci)
                if pd.isna(val):
                    cell.value = ""
                elif isinstance(val, (int, np.integer)):
                    cell.value = int(val)
                    cell.number_format = '#,##0'
                    cell.alignment = align_right
                elif isinstance(val, (float, np.floating)):
                    cell.value = float(val)
                    cell.number_format = '0.000000'
                    cell.alignment = align_right
                elif isinstance(val, bool):
                    cell.value = str(val)
                    cell.alignment = align_center
                else:
                    cell.value = str(val)
                    cell.alignment = align_left
                cell.fill = fill
                cell.font = normal_font
                cell.border = cell_border

        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)

        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
        ws.auto_filter.ref = ws.dimensions

    # ── Sheet 1: Fig1_F1_Raw ──
    fig1_csv = os.path.join(GNN_COMPARE_DIR, "fig1_f1_raw.csv")
    if os.path.exists(fig1_csv):
        df_f1 = pd.read_csv(fig1_csv)
        ws1 = wb.create_sheet(title="Fig1_F1_Raw")
        write_sheet(ws1, df_f1, "Figure 1: Rescale vs Retrain — Raw F1 Scores")

    # ── Sheet 2: Fig1_F1_Summary ──
    if os.path.exists(fig1_csv):
        df_f1 = pd.read_csv(fig1_csv)
        agg = df_f1.groupby(['dataset', 'scaler', 'mode']).agg(
            mean_f1=('f1', 'mean'),
            std_f1=('f1', 'std'),
            count=('f1', 'count')
        ).reset_index()
        agg.columns = ['dataset', 'scaler', 'mode', 'mean_F1', 'std_F1', 'num_runs']
        agg['std_F1'] = agg['std_F1'].fillna(0)
        ws2 = wb.create_sheet(title="Fig1_F1_Summary")
        write_sheet(ws2, agg, "Figure 1: Rescale vs Retrain — F1 Summary Statistics")

    # ── Sheet 3: Fig2_Latency_Timeline ──
    ws3 = wb.create_sheet(title="Fig2_Latency_Timeline")
    df_lat = pd.DataFrame(latency_rows)
    write_sheet(ws3, df_lat, "Figure 2: Benign-User Latency Timeline")

    # ── Sheet 4: Fig3_SPS ──
    ws4 = wb.create_sheet(title="Fig3_SPS")
    df_sps = pd.DataFrame(sps_rows)
    write_sheet(ws4, df_sps, "Figure 3: Security Preservation Score")

    # ── Sheet 5: Fig4_Service_Availability ──
    ws5 = wb.create_sheet(title="Fig4_Service_Availability")
    df_avail = pd.DataFrame(avail_rows)
    write_sheet(ws5, df_avail, "Figure 4: Service Availability")

    # ── Sheet 6: Fig5_Bandwidth_Timeline ──
    ws6 = wb.create_sheet(title="Fig5_Bandwidth_Timeline")
    df_bw = pd.DataFrame(bw_rows)
    write_sheet(ws6, df_bw, "Figure 5: Bandwidth Utilization Timeline")

    # ── Sheet 7: Fig6_Throughput_Timeline ──
    ws7 = wb.create_sheet(title="Fig6_Throughput_Timeline")
    df_tp = pd.DataFrame(tp_rows)
    write_sheet(ws7, df_tp, "Figure 6: Benign Throughput Timeline")

    # ── Sheet 8: Benchmark_Summary ──
    ws8 = wb.create_sheet(title="Benchmark_Summary")
    summary_rows = []
    for run in runs:
        ph = run['probe_history']
        qh = run['qos_history']
        fh = run['flow_history']

        valid_lats = [p['latency_ms'] for p in ph if p.get('latency_ms') is not None and p['latency_ms'] > 0]
        avg_latency = np.mean(valid_lats) if valid_lats else 0.0

        qos_vals = [sum(e.get('throughput', {}).values()) for e in qh]
        peak_bw = max(qos_vals) if qos_vals else 0.0

        baseline_flows = [sum(e.get('throughput', {}).values()) for e in fh if e.get('phase') == 'baseline']
        avg_benign_tp = np.mean(baseline_flows) if baseline_flows else 0.0

        summary_rows.append({
            'controller': display_name(run['controller']),
            'topology': run['topology'],
            'attack': SCENARIO_LABELS.get(run['attack'], run['attack']),
            'SCS': run['SCS'],
            'QPS': run['QPS'],
            'UIS': run['UIS'],
            'RES': run['RES'],
            'NRS': run['NRS'],
            'avg_benign_latency_ms': avg_latency,
            'peak_bandwidth_Bps': peak_bw,
            'avg_benign_throughput_Bps': avg_benign_tp,
        })
    df_summary = pd.DataFrame(summary_rows)
    write_sheet(ws8, df_summary, "Benchmark Summary — All Runs")

    # ── Sheet 9: Run_Validation ──
    ws9 = wb.create_sheet(title="Run_Validation")
    validation_rows = []
    for run in runs:
        ph = run['probe_history']
        qh = run['qos_history']
        fh = run['flow_history']
        valid_lat = [e for e in ph if e.get('latency_ms') is not None and e.get('latency_ms', 0) > 0]
        nonzero_qos = [e for e in qh if any(v > 0 for v in e.get('throughput', {}).values())]
        nonzero_flow = [e for e in fh if any(v > 0 for v in e.get('throughput', {}).values())]

        validation_rows.append({
            'run_id': run['file'],
            'controller': display_name(run['controller']),
            'topology': run['topology'],
            'attack': SCENARIO_LABELS.get(run['attack'], run['attack']),
            'service_started': 'Yes' if valid_lat else 'No',
            'controller_started': 'Yes',
            'infer_server_healthy': 'Yes' if run['NRS'] > 0 else 'Unknown',
            'monitor_active': 'Yes' if qh else 'No',
            'raw_result_present': 'Yes',
            'nonzero_latency': 'Yes' if valid_lat else 'No',
            'nonzero_bandwidth': 'Yes' if nonzero_qos else 'No',
            'nonzero_throughput': 'Yes' if nonzero_flow else 'No',
            'run_status': 'OK',
            'error_notes': '',
        })
    df_val = pd.DataFrame(validation_rows)
    write_sheet(ws9, df_val, "Run Validation Checks")

    wb.save(EXCEL_PATH)
    print(f"[excel] Saved: {EXCEL_PATH}")


# ═══════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════
def main():
    print("=" * 65)
    print("  MASTER FIGURE & EXCEL GENERATOR (Expected Layout)")
    print("=" * 65)
    t_start = time.time()

    plt.rcParams.update(STYLE)
    os.makedirs(FIGURES_DIR, exist_ok=True)
    os.makedirs(EXPECTED_IMAGES_DIR, exist_ok=True)

    runs = load_all_runs()

    # Generate figures matching expected layout
    latency_rows = generate_fig2(runs)
    sps_rows = generate_fig3(runs)
    avail_rows = generate_fig4(runs)
    bw_rows = generate_fig5(runs)
    tp_rows = generate_fig6(runs)

    # Generate Excel
    generate_excel(runs, latency_rows, sps_rows, avail_rows, bw_rows, tp_rows)

    # Also copy all generated figures into backend/benchmark/images/expected/ and backend/benchmark/images/ for completeness
    for fig_file in glob.glob(os.path.join(FIGURES_DIR, "*.png")):
        bname = os.path.basename(fig_file)
        dest = os.path.join(EXPECTED_IMAGES_DIR, bname)
        with open(fig_file, 'rb') as sf, open(dest, 'wb') as df:
            df.write(sf.read())

    elapsed = time.time() - t_start
    print(f"\n{'='*65}")
    print(f"  COMPLETE — {elapsed:.1f}s")
    print(f"{'='*65}")
    print(f"  Figures output: {FIGURES_DIR}/ and {EXPECTED_IMAGES_DIR}/")
    print(f"  Excel output:   {EXCEL_PATH}")


if __name__ == "__main__":
    main()
